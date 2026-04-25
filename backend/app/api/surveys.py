"""
surveys.py — Router FastAPI para encuestas de satisfacción.

CAMBIOS v2 (Lote visual):
  - por_departamento ahora incluye las 5 dimensiones promediadas
    (efficiency, communication, technical_quality, added_value, global_experience)
    para alimentar el heatmap del dashboard.
  - Añadida función helper _avg_dim_for_surveys para evitar repetición.

Endpoints:
    GET  /surveys/kpis            — KPIs del dashboard
    GET  /surveys/                — Listar con filtros y paginación
    GET  /surveys/{id}            — Detalle
    POST /surveys/                — Crear (admin)
    PUT  /surveys/{id}            — Editar (admin)
    DELETE /surveys/{id}          — Eliminar (admin)
    POST /surveys/import          — Importar Excel
"""

import io
import logging
from decimal import Decimal
from math import ceil
from typing import Optional

import openpyxl
from fastapi import (
    APIRouter, Depends, File, HTTPException,
    Query, UploadFile, status,
)
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.dependencies import get_current_user, require_admin
from app.models.survey_models import Survey
from app.models.user_models import User
from app.schemas.survey_schemas import (
    DimensionKPI,
    SurveyCreate,
    SurveyDashboardKPI,
    SurveyImportResponse,
    SurveyKPIPorPeriodo,
    SurveyKPIPorSede,
    SurveyListResponse,
    SurveyResponse,
    SurveyUpdate,
)
from app.services.survey_service import importar_surveys_desde_excel

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/surveys",
    tags=["Encuestas de Satisfacción"],
    responses={404: {"description": "Encuesta no encontrada"}},
)

# Dimensiones con sus nombres legibles para el dashboard
DIMENSIONES_MAP = {
    "efficiency":        "Eficiencia",
    "communication":     "Comunicación",
    "technical_quality": "Calidad Técnica",
    "added_value":       "Valor Agregado",
    "global_experience": "Experiencia Global",
}


def _estado_satisfaccion(pct: Optional[float]) -> str:
    """Semáforo con rangos de satisfacción (90/80), diferente a 5S."""
    if pct is None:
        return "Sin datos"
    if pct >= 0.90:
        return "Excelente"
    if pct >= 0.80:
        return "Aceptable"
    return "Crítico"


def _avg_field(surveys: list, field_name: str) -> Optional[float]:
    """Promedio de un campo numérico sobre una lista de surveys."""
    vals = [
        float(getattr(s, field_name))
        for s in surveys
        if getattr(s, field_name, None) is not None
    ]
    return round(sum(vals) / len(vals), 4) if vals else None


def _get_survey_or_404(survey_id: int, db: Session) -> Survey:
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Encuesta id={survey_id} no encontrada.",
        )
    return survey


def _apply_filters(q, survey_type, department, site, year, quarter, period):
    """Aplica filtros opcionales a la query."""
    if survey_type:
        q = q.filter(Survey.survey_type.ilike(f"%{survey_type}%"))
    if department:
        q = q.filter(Survey.department.ilike(f"%{department}%"))
    if site:
        q = q.filter(Survey.site.ilike(f"%{site}%"))
    if year:
        q = q.filter(Survey.year == year)
    if quarter is not None:
        q = q.filter(Survey.quarter == quarter)
    if period:
        q = q.filter(Survey.period == period)
    return q


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTACIÓN EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def _sfill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

_S_FILL_HDR  = _sfill("B4427F")
_S_FILL_EXC  = _sfill("C6EFCE")
_S_FILL_ACE  = _sfill("FFEB9C")
_S_FILL_CRI  = _sfill("FFC7CE")
_S_FONT_HDR  = Font(bold=True, color="FFFFFF")
_S_ALIGN_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _survey_estado_fill(pct_0_1: Optional[float]) -> PatternFill:
    if pct_0_1 is None:   return _sfill("F2F2F2")
    if pct_0_1 >= 0.90:   return _S_FILL_EXC
    if pct_0_1 >= 0.80:   return _S_FILL_ACE
    return _S_FILL_CRI

def _survey_estado_label(pct_0_1: Optional[float]) -> str:
    if pct_0_1 is None:   return "Sin datos"
    if pct_0_1 >= 0.90:   return "Excelente"
    if pct_0_1 >= 0.80:   return "Aceptable"
    return "Crítico"


@router.get("/export", summary="Exportar encuestas de satisfacción a Excel")
def export_surveys(
    year:         Optional[int] = Query(None, ge=2000, le=2100),
    quarter:      Optional[str] = Query(None, description="Q1, Q2, Q3, Q4"),
    site:         Optional[str] = Query(None),
    survey_type:  Optional[str] = Query(None),
    current_user: User          = Depends(get_current_user),
    db:           Session       = Depends(get_db),
):
    quarter_num = None
    if quarter:
        q_clean = quarter.strip().upper()
        if q_clean.startswith("Q") and len(q_clean) == 2:
            try:
                quarter_num = int(q_clean[1])
            except ValueError:
                pass

    q = db.query(Survey)
    q = _apply_filters(q, survey_type, None, site, year, quarter_num, None)
    surveys = q.order_by(Survey.year.desc(), Survey.quarter.desc(), Survey.department).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Satisfacción"

    headers = [
        "Tipo", "Departamento", "Área", "Sede", "Período", "Año", "Trimestre",
        "Eficiencia", "Comunicación", "Cal. Técnica", "Valor Agregado", "Exp. Global",
        "Sat. Interna", "Sat. Externa", "Estado",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _S_FILL_HDR
        cell.font = _S_FONT_HDR
        cell.alignment = _S_ALIGN_CTR

    for s in surveys:
        sat_i = float(s.internal_satisfaction) if s.internal_satisfaction is not None else None
        sat_e = float(s.external_satisfaction) if s.external_satisfaction is not None else None
        avg   = (sat_i + sat_e) / 2 if sat_i is not None and sat_e is not None else (sat_i or sat_e)
        ws.append([
            s.survey_type,
            s.department or "",
            s.area or "",
            s.site or "",
            s.period_name or s.period or "",
            s.year,
            f"Q{s.quarter}" if s.quarter else "",
            round(float(s.efficiency        or 0) * 100, 1),
            round(float(s.communication     or 0) * 100, 1),
            round(float(s.technical_quality or 0) * 100, 1),
            round(float(s.added_value       or 0) * 100, 1),
            round(float(s.global_experience or 0) * 100, 1),
            round(sat_i * 100, 1) if sat_i is not None else None,
            round(sat_e * 100, 1) if sat_e is not None else None,
            _survey_estado_label(avg),
        ])
        rn = ws.max_row
        fill = _survey_estado_fill(avg)
        ws.cell(rn, 13).fill = fill
        ws.cell(rn, 14).fill = fill
        ws.cell(rn, 15).fill = fill

    col_widths = {
        "A": 18, "B": 26, "C": 22, "D": 18, "E": 16, "F": 8, "G": 10,
        "H": 13, "I": 14, "J": 14, "K": 15, "L": 14,
        "M": 13, "N": 13, "O": 13,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=satisfaccion.xlsx"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# KPIs DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/kpis",
    response_model=SurveyDashboardKPI,
    summary="KPIs del dashboard de satisfacción",
    description=(
        "Devuelve KPIs globales + desglose por período, sede y departamento. "
        "A partir de la v2, `por_departamento` incluye las 5 dimensiones "
        "promediadas para alimentar el heatmap del dashboard."
    ),
)
def get_survey_kpis(
    year:         Optional[int] = Query(None, ge=2000, le=2100),
    quarter:      Optional[str] = Query(None, description="Q1, Q2, Q3, Q4"),
    survey_type:  Optional[str] = Query(None, description="Filtrar por tipo/departamento"),
    site:         Optional[str] = Query(None, description="Filtrar por sede"),
    current_user: User          = Depends(get_current_user),
    db:           Session       = Depends(get_db),
):
    # Convertir quarter string → número
    quarter_num = None
    if quarter:
        q_clean = quarter.strip().upper()
        if q_clean.startswith("Q") and len(q_clean) == 2:
            try:
                quarter_num = int(q_clean[1])
            except ValueError:
                quarter_num = None
        else:
            try:
                quarter_num = int(quarter)
            except ValueError:
                quarter_num = None

    q = db.query(Survey)
    q = _apply_filters(q, survey_type, None, site, year, quarter_num, None)
    surveys = q.all()

    if not surveys:
        return SurveyDashboardKPI(
            total_registros=0,
            periodos_disponibles=[],
        )

    # ── Promedios globales ────────────────────────────────────────────────────
    sat_int = _avg_field(surveys, "internal_satisfaction")
    sat_ext = _avg_field(surveys, "external_satisfaction")

    overall = None
    if sat_int is not None and sat_ext is not None:
        overall = round((sat_int + sat_ext) / 2, 4)
    elif sat_int is not None:
        overall = sat_int
    elif sat_ext is not None:
        overall = sat_ext

    # ── Dimensiones globales ──────────────────────────────────────────────────
    dimensiones = []
    for field_name, nombre in DIMENSIONES_MAP.items():
        prom = _avg_field(surveys, field_name)
        if prom is not None:
            dimensiones.append(DimensionKPI(
                nombre=nombre,
                promedio=prom,
                estado=_estado_satisfaccion(prom),
            ))

    mejor_dim = max(dimensiones, key=lambda d: d.promedio).nombre if dimensiones else None
    peor_dim  = min(dimensiones, key=lambda d: d.promedio).nombre if dimensiones else None

    # ── Por período ───────────────────────────────────────────────────────────
    periodos_dict: dict[str, list] = {}
    for s in surveys:
        key = s.period or "Sin período"
        periodos_dict.setdefault(key, []).append(s)

    por_periodo = []
    for period_key, period_surveys in sorted(periodos_dict.items()):
        s0       = period_surveys[0]
        int_vals = [float(s.internal_satisfaction) for s in period_surveys if s.internal_satisfaction]
        ext_vals = [float(s.external_satisfaction) for s in period_surveys if s.external_satisfaction]
        por_periodo.append(SurveyKPIPorPeriodo(
            period=period_key,
            period_name=s0.period_name or period_key,
            year=s0.year or 0,
            quarter=s0.quarter or 0,
            sat_interna=round(sum(int_vals) / len(int_vals), 4) if int_vals else None,
            sat_externa=round(sum(ext_vals) / len(ext_vals), 4) if ext_vals else None,
            n_registros=len(period_surveys),
        ))

    # ── Por sede ──────────────────────────────────────────────────────────────
    sedes_dict: dict[str, list] = {}
    for s in surveys:
        key = s.site or "Sin sede"
        sedes_dict.setdefault(key, []).append(s)

    por_sede = []
    for sede, sede_surveys in sorted(sedes_dict.items()):
        int_vals = [float(s.internal_satisfaction) for s in sede_surveys if s.internal_satisfaction]
        ext_vals = [float(s.external_satisfaction) for s in sede_surveys if s.external_satisfaction]
        por_sede.append(SurveyKPIPorSede(
            site=sede,
            sat_interna=round(sum(int_vals) / len(int_vals), 4) if int_vals else None,
            sat_externa=round(sum(ext_vals) / len(ext_vals), 4) if ext_vals else None,
            n_registros=len(sede_surveys),
        ))

    # ── Por departamento (v2: incluye las 5 dimensiones) ─────────────────────
    dept_dict: dict[str, list] = {}
    for s in surveys:
        dept_dict.setdefault(s.department, []).append(s)

    por_departamento = []
    for dept, dept_surveys in sorted(dept_dict.items()):
        int_vals = [float(s.internal_satisfaction) for s in dept_surveys if s.internal_satisfaction]
        ext_vals = [float(s.external_satisfaction) for s in dept_surveys if s.external_satisfaction]

        sat_i = round(sum(int_vals) / len(int_vals), 4) if int_vals else None
        sat_e = round(sum(ext_vals) / len(ext_vals), 4) if ext_vals else None

        por_departamento.append({
            # Identificación
            "departamento": dept,
            "n_registros":  len(dept_surveys),
            # Satisfacción
            "sat_interna":  sat_i,
            "sat_externa":  sat_e,
            "estado":       _estado_satisfaccion(sat_i),
            # ── v2: 5 dimensiones promediadas por departamento ────────────────
            # Estos campos alimentan el heatmap del dashboard.
            # Si el departamento no tiene datos de una dimensión, se devuelve None.
            "efficiency":        _avg_field(dept_surveys, "efficiency"),
            "communication":     _avg_field(dept_surveys, "communication"),
            "technical_quality": _avg_field(dept_surveys, "technical_quality"),
            "added_value":       _avg_field(dept_surveys, "added_value"),
            "global_experience": _avg_field(dept_surveys, "global_experience"),
        })

    return SurveyDashboardKPI(
        total_registros=len(surveys),
        periodos_disponibles=sorted(periodos_dict.keys()),
        sat_interna_global=sat_int,
        sat_externa_global=sat_ext,
        overall_global=overall,
        dimensiones=dimensiones,
        mejor_dimension=mejor_dim,
        peor_dimension=peor_dim,
        por_periodo=por_periodo,
        por_sede=por_sede,
        por_departamento=por_departamento,
    )


# ─────────────────────────────────────────────────────────────────────────────
# IMPORTACIÓN (siempre antes de las rutas con {id})
# ─────────────────────────────────────────────────────────────────────────────

@router.post(
    "/import",
    response_model=SurveyImportResponse,
    summary="Importar encuestas desde Excel",
    description=(
        "Acepta el formato exacto de `Satisfaccion_Estructura_Mejorada.xlsx`. "
        "Lee la hoja `Hechos_Satisfaccion` por defecto. "
        "Con `overwrite=True` actualiza duplicados (dept+site+period+type)."
    ),
)
async def import_surveys(
    file:       UploadFile = File(...),
    overwrite:  bool       = Query(False, description="Actualizar duplicados"),
    sheet_name: str        = Query("Hechos_Satisfaccion", description="Hoja del Excel"),
    _:          User       = Depends(require_admin),
    db:         Session    = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos Excel (.xlsx o .xls).")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(400, "El archivo está vacío.")

    logger.info(f"Importando surveys: '{file.filename}' | {len(file_bytes)} bytes")

    result = importar_surveys_desde_excel(
        file_bytes=file_bytes,
        db=db,
        sheet_name=sheet_name,
        overwrite_if_exists=overwrite,
    )

    return SurveyImportResponse(
        message=(
            f"Importación completada: {result.nuevas} nuevas, "
            f"{result.actualizadas} actualizadas, "
            f"{result.omitidas} omitidas."
        ),
        total_filas=result.total_filas,
        nuevas=result.nuevas,
        actualizadas=result.actualizadas,
        omitidas=result.omitidas,
        errores_n=len(result.errores),
        errores=result.errores[:50],
        survey_ids=result.survey_ids,
    )


# ─────────────────────────────────────────────────────────────────────────────
# CRUD
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/",
    response_model=SurveyListResponse,
    summary="Listar encuestas con filtros y paginación",
)
def list_surveys(
    survey_type:  Optional[str] = Query(None),
    department:   Optional[str] = Query(None),
    site:         Optional[str] = Query(None),
    year:         Optional[int] = Query(None, ge=2000, le=2100),
    quarter:      Optional[str] = Query(None, pattern=r"^Q[1-4]$"),
    period:       Optional[str] = Query(None),
    page:         int           = Query(1, ge=1),
    page_size:    int           = Query(20, ge=1, le=100),
    order_by:     str           = Query("department"),
    order_dir:    str           = Query("asc", pattern="^(asc|desc)$"),
    current_user: User          = Depends(get_current_user),
    db:           Session       = Depends(get_db),
):
    # Convertir quarter
    quarter_num = None
    if quarter:
        q_clean = quarter.strip().upper()
        if q_clean.startswith("Q") and len(q_clean) == 2:
            quarter_num = int(q_clean[1])

    q = db.query(Survey)
    q = _apply_filters(q, survey_type, department, site, year, quarter_num, period)

    total = q.count()

    order_map = {
        "department":            Survey.department,
        "site":                  Survey.site,
        "internal_satisfaction": Survey.internal_satisfaction,
        "external_satisfaction": Survey.external_satisfaction,
        "year":                  Survey.year,
        "quarter":               Survey.quarter,
    }
    order_col = order_map.get(order_by, Survey.department)
    q = q.order_by(order_col.desc() if order_dir == "desc" else order_col.asc())

    surveys     = q.offset((page - 1) * page_size).limit(page_size).all()
    total_pages = ceil(total / page_size) if total > 0 else 1

    return SurveyListResponse(
        items=[SurveyResponse.from_orm_with_extras(s) for s in surveys],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
        has_next=page < total_pages,
        has_prev=page > 1,
    )


@router.get("/{survey_id}", response_model=SurveyResponse)
def get_survey(
    survey_id:    int,
    current_user: User    = Depends(get_current_user),
    db:           Session = Depends(get_db),
):
    return SurveyResponse.from_orm_with_extras(_get_survey_or_404(survey_id, db))


@router.post(
    "/",
    response_model=SurveyResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Crear encuesta (admin)",
)
def create_survey(
    survey_in: SurveyCreate,
    _:         User    = Depends(require_admin),
    db:        Session = Depends(get_db),
):
    existing = db.query(Survey).filter(
        Survey.survey_type == survey_in.survey_type,
        Survey.department  == survey_in.department,
        Survey.site        == survey_in.site,
        Survey.period      == survey_in.period,
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"Ya existe una encuesta para tipo='{survey_in.survey_type}', "
                f"dept='{survey_in.department}', sede='{survey_in.site}', "
                f"período='{survey_in.period}'."
            ),
        )

    survey = Survey(
        **{
            k: (Decimal(str(v)) if isinstance(v, float) and v is not None else v)
            for k, v in survey_in.model_dump().items()
        },
        import_source="manual",
    )
    db.add(survey)
    db.commit()
    db.refresh(survey)
    return SurveyResponse.from_orm_with_extras(survey)


@router.put("/{survey_id}", response_model=SurveyResponse, summary="Editar encuesta (admin)")
def update_survey(
    survey_id: int,
    survey_in: SurveyUpdate,
    _:         User    = Depends(require_admin),
    db:        Session = Depends(get_db),
):
    survey = _get_survey_or_404(survey_id, db)

    for field_name, value in survey_in.model_dump(exclude_unset=True).items():
        if isinstance(value, float) and value is not None:
            setattr(survey, field_name, Decimal(str(value)))
        else:
            setattr(survey, field_name, value)

    db.commit()
    db.refresh(survey)
    return SurveyResponse.from_orm_with_extras(survey)


@router.delete(
    "/{survey_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Eliminar encuesta (admin)",
)
def delete_survey(
    survey_id: int,
    _:         User    = Depends(require_admin),
    db:        Session = Depends(get_db),
):
    survey = _get_survey_or_404(survey_id, db)
    db.delete(survey)
    db.commit()