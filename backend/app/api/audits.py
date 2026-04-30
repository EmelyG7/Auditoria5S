"""
audits.py — Router FastAPI para el módulo de Auditorías 5S.

Endpoints:
    POST   /audits/              — Crear auditoría desde formulario web
    GET    /audits/              — Listar con paginación y filtros
    GET    /audits/types         — Listar tipos de auditoría (para selects del frontend)
    GET    /audits/kpis          — KPIs agregados para el dashboard
    GET    /audits/{id}          — Detalle completo de una auditoría
    PUT    /audits/{id}          — Editar auditoría
    DELETE /audits/{id}          — Eliminar auditoría
    POST   /audits/import        — Importar masivamente desde Excel

NOTA: Los endpoints de dashboard (KPIs, gráficas) están aquí temporalmente.
En una versión futura se moverán a app/api/reports.py para separar
la lógica de reporting de la de CRUD.

NOTA AUTENTICACIÓN: Los comentarios `# TODO: auth` marcan los puntos
donde se añadirá `current_user: User = Depends(get_current_user)`
cuando implementemos JWT en el siguiente paso.
"""

import io
import logging
from math import ceil
from typing import Optional

import openpyxl
from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import extract

# Añadir junto a los imports existentes de fastapi
from app.core.dependencies import get_current_user, require_admin
from app.models.user_models import User

from sqlalchemy import extract, func
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.audit_models import Audit, AuditQuestion, AuditType
from app.schemas.audit_schemas import (
    AuditCreate,
    AuditDashboardKPI,
    AuditDetailResponse,
    AuditFilters,
    AuditImportResponse,
    AuditKPISucursal,
    AuditListResponse,
    AuditResponse,
    AuditTypeResponse,
    AuditUpdate,
    PuntajesPorS,
)
from app.services.audit_service import (
    GrupoS,
    calcular_puntajes_desde_dict,
    crear_audit_desde_calculo,
    get_audit_resumen,
    importar_desde_excel,
    _semaforo,
    AuditCalculationResult,
)
from app.services.audit_analysis_service import analyze_audit, analyze_branch_trend

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/audits",
    tags=["Auditorías 5S"],
    responses={
        404: {"description": "Auditoría no encontrada"},
        422: {"description": "Error de validación en los datos enviados"},
    },
)


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS INTERNOS
# ─────────────────────────────────────────────────────────────────────────────

def _get_audit_or_404(audit_id: int, db: Session) -> Audit:
    """Retorna la auditoría o lanza 404."""
    audit = db.query(Audit).filter(Audit.id == audit_id).first()
    if not audit:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Auditoría con id={audit_id} no encontrada.",
        )
    return audit


def _build_query_with_filters(db: Session, filters: AuditFilters):
    """Construye la query SQLAlchemy aplicando todos los filtros activos."""
    q = db.query(Audit)

    if filters.audit_type_id is not None:
        q = q.filter(Audit.audit_type_id == filters.audit_type_id)

    if filters.branch:
        q = q.filter(Audit.branch == filters.branch)   # <-- CAMBIO: exacto, no like

    if filters.status:
        q = q.filter(Audit.status == filters.status)

    if filters.year:
        q = q.filter(extract("year", Audit.audit_date) == filters.year)

    if filters.quarter:
        # quarter debe ser un número 1-4
        q_num = filters.quarter
        if isinstance(filters.quarter, str) and filters.quarter.startswith("Q"):
            q_num = int(filters.quarter[1])
        else:
            q_num = int(filters.quarter)
        month_start = (q_num - 1) * 3 + 1
        month_end   = q_num * 3
        q = q.filter(
            extract("month", Audit.audit_date) >= month_start,
            extract("month", Audit.audit_date) <= month_end,
        )

    if filters.date_from:
        q = q.filter(Audit.audit_date >= filters.date_from)

    if filters.date_to:
        q = q.filter(Audit.audit_date <= filters.date_to)

    if filters.auditor_email:
        q = q.filter(Audit.auditor_email.ilike(f"%{filters.auditor_email}%"))

    return q


def _grupos_desde_preguntas(questions_create) -> list[GrupoS]:
    """
    Reconstruye la lista de GrupoS a partir de las preguntas del formulario.
    Agrupa por s_index y ordena por question_order.
    """
    grupos_dict: dict[int, GrupoS] = {}
    for q in questions_create:
        if q.s_index not in grupos_dict:
            grupos_dict[q.s_index] = GrupoS(
                s_index=q.s_index,
                nombre_s=q.s_name,
                columnas_preguntas=[],
                columna_observacion=None,
            )
        # Usamos el texto de la pregunta como clave (con el peso embebido para compatibilidad)
        col_key = f"{q.question_text} {q.weight}%"
        grupos_dict[q.s_index].columnas_preguntas.append(col_key)

    return [grupos_dict[k] for k in sorted(grupos_dict.keys())]


def _respuestas_desde_create(audit_create: AuditCreate) -> dict:
    """
    Convierte un AuditCreate en el dict de respuestas que espera
    calcular_puntajes_desde_dict().
    """
    respuestas: dict = {
        "FechaAuditoria": str(audit_create.audit_date),
        "Sucursal":       audit_create.branch,
        "Auditor":        audit_create.auditor_name or "",
        "Email":          audit_create.auditor_email or "",
        "HoraInicio":     str(audit_create.start_time) if audit_create.start_time else None,
        "HoraFin":        str(audit_create.end_time) if audit_create.end_time else None,
    }
    for q in audit_create.questions:
        col_key = f"{q.question_text} {q.weight}%"
        respuestas[col_key] = q.response_percent
        # Observation: la columna de observación de cada S
        obs_key = f"Observaciones {q.s_index + 1}"
        if q.observation and obs_key not in respuestas:
            respuestas[obs_key] = q.observation

    return respuestas


# ─────────────────────────────────────────────────────────────────────────────
# ENDPOINTS — CATÁLOGO Y KPIS (sin {id} para que no colisionen)
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/types",
    response_model=list[AuditTypeResponse],
    summary="Listar tipos de auditoría",
    description="Retorna el catálogo de tipos (Almacenes, Centro de Servicios, RMA). "
                "Útil para poblar selects en el frontend.",
)
def list_audit_types(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.query(AuditType).order_by(AuditType.id).all()


@router.get(
    "/kpis",
    response_model=AuditDashboardKPI,
    summary="KPIs globales para el dashboard",
    description="Retorna métricas agregadas: promedio global, totales, "
                "mejor/peor sucursal y desglose por S.",
)
def get_dashboard_kpis(
   audit_type_id: Optional[int] = Query(None, description="Filtrar por tipo"),
    year:          Optional[int] = Query(None, ge=2000, le=2100),
    quarter:       Optional[str] = Query(None, pattern=r"^Q[1-4]$", description="Trimestre: Q1, Q2, Q3, Q4"),
    branch:        Optional[str] = Query(None, description="Filtrar por sucursal exacta"),
    current_user:  User = Depends(get_current_user),
    db:            Session = Depends(get_db),
):
    # Construir la query manualmente (sin usar AuditFilters para evitar conflictos de tipos)
    q = db.query(Audit)

    if audit_type_id is not None:
        q = q.filter(Audit.audit_type_id == audit_type_id)

    if year:
        q = q.filter(extract('year', Audit.audit_date) == year)

    if quarter:
        q_num = int(quarter[1])  # "Q1" -> 1, "Q2" -> 2, etc.
        month_start = (q_num - 1) * 3 + 1
        month_end   = q_num * 3
        q = q.filter(
            extract('month', Audit.audit_date) >= month_start,
            extract('month', Audit.audit_date) <= month_end
        )

    if branch:
        q = q.filter(Audit.branch == branch)

    audits = q.all()

    if not audits:
        return AuditDashboardKPI(
            promedio_global=0.0,
            estado_global="Sin datos",
            total_auditorias=0,
            sucursales_cumple_pct=0.0,
            sucursales_critico_pct=0.0,
        )

    pcts = [float(a.percentage or 0) for a in audits]
    promedio_global = round(sum(pcts) / len(pcts), 2)

    # Agrupar por sucursal
    sucursal_pcts: dict[str, list[float]] = {}
    for a in audits:
        sucursal_pcts.setdefault(a.branch, []).append(float(a.percentage or 0))

    por_sucursal_stats = {
        branch: {
            "promedio": round(sum(vals) / len(vals), 2),
            "min":      round(min(vals), 2),
            "max":      round(max(vals), 2),
            "n":        len(vals),
        }
        for branch, vals in sucursal_pcts.items()
    }

    n_sucursales    = len(por_sucursal_stats)
    n_cumple        = sum(1 for s in por_sucursal_stats.values() if s["promedio"] >= 80)
    n_critico       = sum(1 for s in por_sucursal_stats.values() if s["promedio"] < 60)
    mejor_branch    = max(por_sucursal_stats, key=lambda b: por_sucursal_stats[b]["promedio"])
    peor_branch     = min(por_sucursal_stats, key=lambda b: por_sucursal_stats[b]["promedio"])

    # Promedios por S
    def avg_s(field: str) -> float:
        vals = [float(getattr(a, field) or 0) for a in audits]
        return round(sum(vals) / len(vals), 2) if vals else 0.0

    # Por tipo de auditoría
    tipo_pcts: dict[str, list[float]] = {}
    for a in audits:
        tipo_nombre = a.audit_type.name if a.audit_type else str(a.audit_type_id)
        tipo_pcts.setdefault(tipo_nombre, []).append(float(a.percentage or 0))

    por_tipo = [
        {
            "tipo":         tipo,
            "promedio":     round(sum(vals) / len(vals), 2),
            "n_auditorias": len(vals),
            "estado":       _semaforo(round(sum(vals) / len(vals), 2)),
        }
        for tipo, vals in tipo_pcts.items()
    ]

    return AuditDashboardKPI(
        promedio_global=promedio_global,
        estado_global=_semaforo(promedio_global),
        total_auditorias=len(audits),
        sucursales_cumple_pct=round((n_cumple / n_sucursales) * 100, 1) if n_sucursales else 0.0,
        sucursales_critico_pct=round((n_critico / n_sucursales) * 100, 1) if n_sucursales else 0.0,
        mejor_sucursal=mejor_branch,
        mejor_sucursal_pct=por_sucursal_stats[mejor_branch]["promedio"],
        peor_sucursal=peor_branch,
        peor_sucursal_pct=por_sucursal_stats[peor_branch]["promedio"],
        por_tipo=por_tipo,
        por_sucursal=[
            AuditKPISucursal(
                branch=branch,
                promedio_pct=stats["promedio"],
                min_pct=stats["min"],
                max_pct=stats["max"],
                n_auditorias=stats["n"],
                estado=_semaforo(stats["promedio"]),
            )
            for branch, stats in sorted(por_sucursal_stats.items(), key=lambda x: x[1]["promedio"], reverse=True)
        ],
        promedio_por_s=PuntajesPorS(
            seiri=    avg_s("seiri_percentage"),
            seiton=   avg_s("seiton_percentage"),
            seiso=    avg_s("seiso_percentage"),
            seiketsu= avg_s("seiketsu_percentage"),
            shitsuke= avg_s("shitsuke_percentage"),
        ),
    )


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS PARA EXPORTACIÓN EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def _xfill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _set_col_widths(ws, widths: dict) -> None:
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

_FILL_HEADER  = _xfill("0A4F79")
_FILL_CUMPLE  = _xfill("C6EFCE")
_FILL_MEJORAR = _xfill("FFEB9C")
_FILL_CRITICO = _xfill("FFC7CE")
_FONT_HDR     = Font(bold=True, color="FFFFFF")
_FONT_BOLD    = Font(bold=True)
_ALIGN_CTR    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(vertical="center", wrap_text=True)

def _estado_fill(estado: str) -> PatternFill:
    if estado == "Cumple":      return _FILL_CUMPLE
    if estado == "Por mejorar": return _FILL_MEJORAR
    return _FILL_CRITICO

def _build_export_query(db, audit_type_id, year, quarter, branch):
    q = db.query(Audit)
    if audit_type_id is not None:
        q = q.filter(Audit.audit_type_id == audit_type_id)
    if year:
        q = q.filter(extract("year", Audit.audit_date) == year)
    if quarter:
        q_num = int(quarter[1])
        q = q.filter(
            extract("month", Audit.audit_date) >= (q_num - 1) * 3 + 1,
            extract("month", Audit.audit_date) <= q_num * 3,
        )
    if branch:
        q = q.filter(Audit.branch == branch)
    return q.order_by(Audit.audit_date.desc())


# ─────────────────────────────────────────────────────────────────────────────
# ENDPOINTS — EXPORTACIONES EXCEL
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/export/summary", summary="Exportar resumen de auditorías a Excel")
def export_audits_summary(
    audit_type_id: Optional[int] = Query(None),
    year:          Optional[int] = Query(None, ge=2000, le=2100),
    quarter:       Optional[str] = Query(None, pattern=r"^Q[1-4]$"),
    branch:        Optional[str] = Query(None),
    current_user:  User = Depends(get_current_user),
    db:            Session = Depends(get_db),
):
    audits = _build_export_query(db, audit_type_id, year, quarter, branch).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"

    headers = [
        "Fecha", "Sucursal", "Tipo de Auditoría", "Auditor",
        "Seiri (%)", "Seiton (%)", "Seiso (%)", "Seiketsu (%)", "Shitsuke (%)",
        "Total (%)", "Estado",
    ]
    ws.append(headers)
    for col, cell in enumerate(ws[1], 1):
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HDR
        cell.alignment = _ALIGN_CTR

    for a in audits:
        pct    = float(a.percentage or 0)
        estado = _semaforo(pct)
        ws.append([
            a.audit_date,
            a.branch,
            a.audit_type.name if a.audit_type else "",
            a.auditor_name or "",
            round(float(a.seiri_percentage    or 0), 1),
            round(float(a.seiton_percentage   or 0), 1),
            round(float(a.seiso_percentage    or 0), 1),
            round(float(a.seiketsu_percentage or 0), 1),
            round(float(a.shitsuke_percentage or 0), 1),
            round(pct, 1),
            estado,
        ])
        rn = ws.max_row
        ws.cell(rn, 1).number_format = "YYYY-MM-DD"
        fill = _estado_fill(estado)
        ws.cell(rn, 10).fill = fill
        ws.cell(rn, 11).fill = fill

    _set_col_widths(ws, {"A": 14, "B": 22, "C": 24, "D": 22,
                         "E": 11, "F": 11, "G": 11, "H": 13, "I": 13,
                         "J": 11, "K": 14})
    ws.freeze_panes = "A2"

    # Hoja pivot: sucursal × trimestre
    ws2 = wb.create_sheet("Pivot Sucursal")
    ws2.append(["Sucursal", "Q1 (%)", "Q2 (%)", "Q3 (%)", "Q4 (%)", "Promedio (%)"])
    for cell in ws2[1]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HDR
        cell.alignment = _ALIGN_CTR

    pivot: dict[str, dict[str, list[float]]] = {}
    for a in audits:
        month   = a.audit_date.month if a.audit_date else 1
        q_label = f"Q{((month - 1) // 3) + 1}"
        pivot.setdefault(a.branch, {}).setdefault(q_label, [])
        pivot[a.branch][q_label].append(float(a.percentage or 0))

    for b, quarters in sorted(pivot.items()):
        all_vals: list[float] = []
        row = [b]
        for ql in ["Q1", "Q2", "Q3", "Q4"]:
            vals = quarters.get(ql, [])
            avg  = round(sum(vals) / len(vals), 1) if vals else None
            row.append(avg)
            if avg is not None:
                all_vals.append(avg)
        overall = round(sum(all_vals) / len(all_vals), 1) if all_vals else None
        row.append(overall)
        ws2.append(row)
        rn = ws2.max_row
        if overall is not None:
            ws2.cell(rn, 6).fill = _estado_fill(_semaforo(overall))

    _set_col_widths(ws2, {"A": 22, "B": 11, "C": 11, "D": 11, "E": 11, "F": 14})
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=auditoria_resumen.xlsx"},
    )


@router.get("/export/detail", summary="Exportar detalle de preguntas a Excel")
def export_audits_detail(
    audit_type_id: Optional[int] = Query(None),
    year:          Optional[int] = Query(None, ge=2000, le=2100),
    quarter:       Optional[str] = Query(None, pattern=r"^Q[1-4]$"),
    branch:        Optional[str] = Query(None),
    current_user:  User = Depends(get_current_user),
    db:            Session = Depends(get_db),
):
    audits = _build_export_query(db, audit_type_id, year, quarter, branch).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle Preguntas"

    headers = [
        "Fecha", "Sucursal", "Tipo", "S (Categoría)", "Pregunta",
        "Peso (%)", "Respuesta (%)", "Puntos Obtenidos", "Puntos Perdidos",
        "Es Crítica", "Observación",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HDR
        cell.alignment = _ALIGN_CTR

    for a in audits:
        for aq in sorted(a.questions, key=lambda x: (x.s_index, x.question_order)):
            ws.append([
                a.audit_date,
                a.branch,
                a.audit_type.name if a.audit_type else "",
                aq.s_name,
                aq.question_text,
                float(aq.weight          or 0),
                float(aq.response_percent or 0),
                float(aq.points_earned   or 0),
                float(aq.points_lost     or 0),
                "Sí" if aq.is_critical else "No",
                aq.observation or "",
            ])
            rn = ws.max_row
            ws.cell(rn, 1).number_format = "YYYY-MM-DD"
            if aq.is_critical:
                ws.cell(rn, 10).fill = _FILL_CRITICO

    _set_col_widths(ws, {"A": 14, "B": 22, "C": 22, "D": 18, "E": 52,
                         "F": 10, "G": 13, "H": 15, "I": 15, "J": 11, "K": 42})
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=auditoria_detalle.xlsx"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# ENDPOINTS — IMPORTACIÓN (antes del {id} para evitar conflictos de ruta)
# ─────────────────────────────────────────────────────────────────────────────

@router.post(
    "/import",
    response_model=AuditImportResponse,
    status_code=status.HTTP_200_OK,
    summary="Importar auditorías desde Excel",
    description=(
        "Sube un archivo Excel con el mismo formato que los checklists originales. "
        "El sistema detecta automáticamente los grupos (Seiri, Seiton, etc.) "
        "extrayendo los pesos de los encabezados de columna.\n\n"
        "Formatos aceptados: .xlsx, .xls\n\n"
        "Si `overwrite` es True, actualiza registros existentes. "
        "Si es False (por defecto), omite duplicados."
    ),
)
async def import_audits_from_excel(
    file:             UploadFile = File(..., description="Archivo Excel del checklist"),
    audit_type_id:    int        = Query(..., description="ID del tipo (1=Almacenes, 2=Centro, 3=RMA)"),
    overwrite:        bool       = Query(False, description="Si True, actualiza duplicados"),
    _: User = Depends(require_admin),
    db:               Session    = Depends(get_db),
):

    # Validar tipo de archivo
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se aceptan archivos Excel (.xlsx o .xls)",
        )

    # Validar que el audit_type_id existe
    audit_type = db.query(AuditType).filter(AuditType.id == audit_type_id).first()
    if not audit_type:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Tipo de auditoría id={audit_type_id} no existe.",
        )

    # Leer el archivo
    try:
        file_bytes = await file.read()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se pudo leer el archivo: {e}",
        )

    if len(file_bytes) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo está vacío.",
        )

    logger.info(
        f"Importando Excel: '{file.filename}' | "
        f"tipo={audit_type.name} | {len(file_bytes)} bytes | overwrite={overwrite}"
    )

    # Ejecutar importación
    result = importar_desde_excel(
        file_bytes=file_bytes,
        audit_type_id=audit_type_id,
        db=db,
        overwrite_if_exists=overwrite,
    )

    return AuditImportResponse(
        message=(
            f"Importación completada: {result.nuevas} nuevas, "
            f"{result.actualizadas} actualizadas, "
            f"{result.omitidas} omitidas, "
            f"{len(result.errores)} errores."
        ),
        total_filas=result.total_filas,
        nuevas=result.nuevas,
        actualizadas=result.actualizadas,
        omitidas=result.omitidas,
        errores_n=len(result.errores),
        errores=result.errores[:50],  # Máximo 50 errores en la respuesta
        audit_ids=result.audits_creados,
    )


# ─────────────────────────────────────────────────────────────────────────────
# ENDPOINTS — CRUD PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

@router.post(
    "/",
    response_model=AuditDetailResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Crear nueva auditoría",
    description=(
        "Crea una auditoría desde el formulario web. "
        "El backend calcula automáticamente todos los puntajes. "
        "Las respuestas deben ser 0, 50 o 100 por pregunta."
    ),
)
def create_audit(
    audit_in: AuditCreate,
    _: User = Depends(require_admin),
    db:       Session = Depends(get_db),
):
    # TODO: auth

    # Validar que el tipo existe
    audit_type = db.query(AuditType).filter(AuditType.id == audit_in.audit_type_id).first()
    if not audit_type:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Tipo de auditoría id={audit_in.audit_type_id} no existe.",
        )

    # Reconstruir grupos y respuestas desde el payload del formulario
    grupos = _grupos_desde_preguntas(audit_in.questions)
    respuestas = _respuestas_desde_create(audit_in)

    # Calcular puntajes
    try:
        resultado = calcular_puntajes_desde_dict(respuestas, grupos)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(e),
        )

    # Añadir observaciones generales si las hay
    resultado.sucursal = audit_in.branch
    resultado.auditor  = audit_in.auditor_name or ""
    resultado.email    = audit_in.auditor_email or ""

    # Persistir
    try:
        audit = crear_audit_desde_calculo(
            resultado=resultado,
            audit_type_id=audit_in.audit_type_id,
            db=db,
            import_source="manual",
            overwrite_if_exists=False,
        )
        # Guardar observación general si viene en el payload
        if audit_in.general_observations:
            audit.general_observations = audit_in.general_observations
        db.commit()
        db.refresh(audit)
    except ValueError as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=str(e),
        )
    except Exception as e:
        db.rollback()
        logger.error(f"Error al crear auditoría: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al guardar la auditoría.",
        )

    return AuditDetailResponse.from_orm_with_extras(audit)


@router.get(
    "/",
    response_model=AuditListResponse,
    summary="Listar auditorías",
    description=(
        "Retorna auditorías con paginación. "
        "Soporta filtros por tipo, sucursal, estado, año, trimestre y rango de fechas."
    ),
)
def list_audits(
    # Filtros
    audit_type_id: Optional[int]  = Query(None),
    branch:        Optional[str]  = Query(None, description="Búsqueda parcial por nombre de sucursal"),
    status_filter: Optional[str]  = Query(None, alias="status", description="'Cumple', 'Por mejorar', 'Crítico'"),
    year:          Optional[int]  = Query(None, ge=2000, le=2100),
    quarter:       Optional[str]  = Query(None, pattern=r"^Q[1-4]$"),
    date_from:     Optional[str]  = Query(None, description="Fecha inicio YYYY-MM-DD"),
    date_to:       Optional[str]  = Query(None, description="Fecha fin YYYY-MM-DD"),
    auditor_email: Optional[str]  = Query(None),
    # Paginación
    page:          int            = Query(1, ge=1, description="Número de página"),
    page_size:     int            = Query(20, ge=1, le=100, description="Registros por página"),
    # Ordenamiento
    order_by:      str            = Query("audit_date", description="Campo por el que ordenar"),
    order_dir:     str            = Query("desc", pattern="^(asc|desc)$"),
    current_user:  User = Depends(get_current_user),
    db:            Session        = Depends(get_db),
):

    # Parsear fechas
    from datetime import date as date_type
    date_from_parsed = None
    date_to_parsed   = None
    if date_from:
        try:
            date_from_parsed = date_type.fromisoformat(date_from)
        except ValueError:
            raise HTTPException(400, f"date_from inválido: '{date_from}'. Usa YYYY-MM-DD.")
    if date_to:
        try:
            date_to_parsed = date_type.fromisoformat(date_to)
        except ValueError:
            raise HTTPException(400, f"date_to inválido: '{date_to}'. Usa YYYY-MM-DD.")

    filters = AuditFilters(
        audit_type_id=audit_type_id,
        branch=branch,
        status=status_filter,
        year=year,
        quarter=quarter,
        date_from=date_from_parsed,
        date_to=date_to_parsed,
        auditor_email=auditor_email,
    )

    q = _build_query_with_filters(db, filters)

    # Total antes de paginar
    total = q.count()

    # Ordenamiento
    order_col_map = {
        "audit_date": Audit.audit_date,
        "branch":     Audit.branch,
        "percentage": Audit.percentage,
        "status":     Audit.status,
        "created_at": Audit.created_at,
    }
    order_col = order_col_map.get(order_by, Audit.audit_date)
    if order_dir == "desc":
        q = q.order_by(order_col.desc())
    else:
        q = q.order_by(order_col.asc())

    # Paginación
    offset    = (page - 1) * page_size
    audits    = q.offset(offset).limit(page_size).all()
    total_pages = ceil(total / page_size) if total > 0 else 1

    return AuditListResponse(
        items=[AuditResponse.from_orm_with_extras(a) for a in audits],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
        has_next=page < total_pages,
        has_prev=page > 1,
    )


@router.get(
    "/branch-trend",
    summary="Tendencia histórica de una sucursal",
    description=(
        "Retorna la evolución de puntajes (global y por S) de una sucursal "
        "en un tipo de auditoría a lo largo del tiempo. "
        "Útil para el gráfico de líneas del análisis."
    ),
)
def get_branch_trend(
    branch:        str     = Query(..., description="Nombre de la sucursal"),
    audit_type_id: int     = Query(..., description="ID del tipo de auditoría"),
    limit:         int     = Query(10, ge=2, le=30, description="Máximo de auditorías a incluir"),
    current_user:  User    = Depends(get_current_user),
    db:            Session = Depends(get_db),
):
    return analyze_branch_trend(
        branch=branch,
        audit_type_id=audit_type_id,
        db=db,
        limit=limit,
    )


@router.get(
    "/{audit_id}/analysis",
    summary="Análisis inteligente de una auditoría",
    description=(
        "Genera un análisis completo de la auditoría comparándola con las "
        "anteriores de la misma sucursal y tipo:\n\n"
        "- **vs_previous**: comparativa con la auditoría inmediata anterior\n"
        "- **s_analysis**: análisis por cada S (tendencia, delta, estancamiento)\n"
        "- **stagnant_s**: S que no mejoran en varias auditorías consecutivas\n"
        "- **critical_questions**: preguntas con 0% de cumplimiento\n"
        "- **recurrent_findings**: hallazgos que aparecen en múltiples auditorías\n"
        "- **comment_topics**: temas más frecuentes en los comentarios\n"
        "- **executive_summary**: párrafo narrativo generado automáticamente\n"
        "- **recommendations**: acciones sugeridas priorizadas"
    ),
)
def get_audit_analysis(
    audit_id:     int,
    history_n:    int     = Query(5, ge=2, le=20, description="Auditorías anteriores a considerar"),
    current_user: User    = Depends(get_current_user),
    db:           Session = Depends(get_db),
):
    result = analyze_audit(audit_id=audit_id, db=db, history_n=history_n)
    if "error" in result:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=result["error"])
    return result


@router.get(
    "/{audit_id}",
    response_model=AuditDetailResponse,
    summary="Detalle de una auditoría",
    description="Retorna la auditoría completa con todas las preguntas, "
                "puntajes por S y lista de preguntas críticas.",
)
def get_audit(
    audit_id: int,
    current_user: User = Depends(get_current_user),
    db:       Session = Depends(get_db),
):
    audit = _get_audit_or_404(audit_id, db)
    return AuditDetailResponse.from_orm_with_extras(audit)


@router.put(
    "/{audit_id}",
    response_model=AuditDetailResponse,
    summary="Editar auditoría",
    description=(
        "Actualiza los campos enviados. "
        "Si se incluye 'questions', recalcula todos los puntajes. "
        "Si no se incluye 'questions', solo actualiza metadatos."
    ),
)
def update_audit(
    audit_id:  int,
    audit_in:  AuditUpdate,
    _: User = Depends(require_admin),
    db:        Session = Depends(get_db),
):
    audit = _get_audit_or_404(audit_id, db)

    # ── Actualizar metadatos simples ──────────────────────────────────────────
    simple_fields = [
        "audit_date", "branch", "auditor_name", "auditor_email",
        "start_time", "end_time", "general_observations",
    ]
    for field_name in simple_fields:
        value = getattr(audit_in, field_name)
        if value is not None:
            setattr(audit, field_name, value)

    # ── Si viene questions, recalcular puntajes ───────────────────────────────
    if audit_in.questions is not None:
        grupos = _grupos_desde_preguntas(audit_in.questions)
        respuestas = {
            "FechaAuditoria": str(audit.audit_date),
            "Sucursal":       audit.branch,
            "Auditor":        audit.auditor_name or "",
            "Email":          audit.auditor_email or "",
        }
        for q in audit_in.questions:
            col_key = f"{q.question_text} {q.weight}%"
            respuestas[col_key] = q.response_percent
            obs_key = f"Observaciones {q.s_index + 1}"
            if q.observation and obs_key not in respuestas:
                respuestas[obs_key] = q.observation

        try:
            resultado = calcular_puntajes_desde_dict(respuestas, grupos)
        except ValueError as e:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=str(e),
            )

        # Actualizar puntajes en el objeto audit
        from decimal import Decimal
        audit.total_score           = Decimal(str(resultado.puntaje_total))
        audit.max_score             = Decimal(str(resultado.puntaje_maximo))
        audit.percentage            = Decimal(str(resultado.porcentaje_general))
        audit.status                = resultado.estado_general

        from app.services.audit_service import S_TO_AUDIT_FIELD
        for ps in resultado.puntajes_por_s:
            campo = S_TO_AUDIT_FIELD.get(ps.nombre_s)
            if campo:
                setattr(audit, campo, Decimal(str(ps.porcentaje)))

        # Reemplazar preguntas
        db.query(AuditQuestion).filter(
            AuditQuestion.audit_id == audit_id
        ).delete(synchronize_session="fetch")
        db.flush()

        from app.services.audit_service import AuditCalculationResult
        for ps in resultado.puntajes_por_s:
            for preg in ps.preguntas:
                aq = AuditQuestion(
                    audit_id        = audit_id,
                    s_name          = ps.nombre_s,
                    s_index         = ps.s_index,
                    question_text   = preg["texto"],
                    question_order  = preg["orden"],
                    weight          = Decimal(str(preg["peso"])),
                    response_percent= Decimal(str(preg["respuesta_pct"])),
                    points_earned   = Decimal(str(preg["puntos"])),
                    observation     = ps.observacion,
                    is_critical     = preg["es_critica"],
                    points_lost     = Decimal(str(preg["puntos_perdidos"])),
                )
                db.add(aq)

    try:
        db.commit()
        db.refresh(audit)
    except Exception as e:
        db.rollback()
        logger.error(f"Error al actualizar auditoría {audit_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al actualizar la auditoría.",
        )

    return AuditDetailResponse.from_orm_with_extras(audit)


@router.delete(
    "/{audit_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Eliminar auditoría",
    description=(
        "Elimina permanentemente la auditoría y todas sus preguntas (CASCADE). "
        "Esta operación no se puede deshacer."
    ),
)
def delete_audit(
    audit_id: int,
    _: User = Depends(require_admin),
    db:       Session = Depends(get_db),
):
    audit = _get_audit_or_404(audit_id, db)

    try:
        db.delete(audit)
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Error al eliminar auditoría {audit_id}: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error interno al eliminar la auditoría.",
        )

    logger.info(f"Auditoría id={audit_id} eliminada.")
    # 204 No Content no retorna body